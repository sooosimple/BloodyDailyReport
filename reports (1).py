# -*- coding: utf-8 -*-
import datetime
import os
import openpyxl as xl
import pandas as pd
from numpy import arange
from calendar import monthrange
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import IconSetRule, ColorScaleRule
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

class ReportGeneral:

    def __init__(self, name):

        if not os.path.isdir('{}'.format(name)):
            os.mkdir('{}'.format(name))
        os.chdir('{}'.format(name))
        if not os.path.isdir('{}Reports'.format(name[:4])):
            os.mkdir('{}Reports'.format(name[:4]))
        if not os.path.isdir('{}Data'.format(name[:4])):
            os.mkdir('{}Data'.format(name[:4]))
        os.chdir("..")

        self.directory_main_script = os.getcwd()
        self.directory_data = '{}Data'.format(name[:4])
        self.directory_report = '{}Reports'.format(name[:4])
        self.report_name = name
        self._date_report = datetime.date.today() - datetime.timedelta(days=1)
        self._date_report_last_day_month = monthrange(self._date_report.year, self._date_report.month)[1]
        self._date_report_week = self._date_report.isocalendar().week
        self._data_book_directory = '{}\\{}\\{}'.format(self.directory_main_script, self.report_name,
                                                        self.directory_data)
        self._report_book_directory = '{}\\{}\\{}'.format(self.directory_main_script, self.report_name,
                                                          self.directory_report)
        self._service_sheet = None
        self.report = None

    def __str__(self):
        return '{} \n Дата отчета: {} \n Директория для данных: {} \n Директория сохранения отчета: {}'. \
            format(self.report_name, self._date_report, self._data_book_directory, self._report_book_directory)

    def replacing_directory(self, directory_data, directory_report):
        self._data_book_directory = directory_data
        self._report_book_directory = directory_report

    def replacing_date(self, date):
        self._date_report = datetime.datetime.strptime(date, '%Y-%m-%d').date()
        self._date_report_last_day_month = monthrange(self._date_report.year, self._date_report.month)[1]
        self._date_report_week = self._date_report.isocalendar().week

    def reading_data_csv(self, book_directory, book_name, sep=','):
        print('Идет загрузка файл {}\\{}.csv'.format(book_directory, book_name))
        list_csv = pd.read_csv('{}\\{}.csv'.format(book_directory, book_name), sep=sep)
        print('Данные загружены')
        print('______________________________________________________________')
        return list_csv

    def reading_data_xlsx(self, book_directory, book_name, name_list='Лист1'):
        print('Идет загрузка листа: {}'.format(name_list))
        list_xl = pd.read_excel('{}\\{}.xlsx'.format(book_directory, book_name), sheet_name=name_list)
        print('Лист загружен')
        print('______________________________________________________________')
        return list_xl

    def save_report(self, report, name_sheet='report', startrow=0, startcol=0, mode='a', overlay=None, print_info=1):
        with pd.ExcelWriter('{}\\{}_{}.xlsx'.format(self._report_book_directory, self.report_name, self._date_report),
                            mode=mode, engine='openpyxl', if_sheet_exists=overlay) as writer:
            report.to_excel(writer, sheet_name=name_sheet, startrow=startrow, startcol=startcol)
        if print_info == 1:
            print('Добавил в {}\\{}_{}.xlsx лист с данными "{}"'.
                  format(self._report_book_directory, self.report_name, self._date_report, name_sheet))
            print('______________________________________________________________')


class DailyReportOpti(ReportGeneral):

    def __init__(self):
        super().__init__(name='Daily_report_OPTI')

        if not os.path.isdir('diagrams_service_folder'):
            os.mkdir('Diagrams_service_folder')

        self.list_spill = None
        self.list_info_azs = None
        self.list_receipts = None
        self.list_price = None
        self.list_last_year = None
        self.list_current_plan = None
        self.list_number_of_azs = None
        self.pumping_gas_stations = None
        self.directory_chart = '{}/Diagrams_service_folder'.format(self.directory_main_script)

    def __str__(self):
        return super().__str__()

    def checking_data_file_directory(self):
        return os.path.isfile('{}/data.xlsx'.format(self._data_book_directory))

    def reading_workbook_xlsx(self):
        self.list_spill = super().reading_data_xlsx(book_directory=self._data_book_directory, book_name='data',
                                                    name_list='spill')
        self.list_info_azs = super().reading_data_xlsx(book_directory=self._data_book_directory, book_name='data',
                                                       name_list='info_AZS')
        self.list_receipts = super().reading_data_xlsx(book_directory=self._data_book_directory, book_name='data',
                                                       name_list='receipts')
        self.list_price = super().reading_data_xlsx(book_directory=self._data_book_directory, book_name='data',
                                                    name_list='price')
        self.list_last_year = super().reading_data_xlsx(book_directory=self._data_book_directory, book_name='data',
                                                        name_list='last_year')
        self.list_current_plan = super().reading_data_xlsx(book_directory=self._data_book_directory, book_name='data',
                                                           name_list='сurrent_plan')
        self.list_number_of_azs = super().reading_data_xlsx(book_directory=self._data_book_directory, book_name='data',
                                                            name_list='number_of_AZS')
        self.list_sales_for_10_minutes = super().reading_data_xlsx(book_directory=self._data_book_directory,
                                                                   book_name='Sales by 10-minutes Report',
                                                                   name_list='Sales by 10-minutes Report')

        self._service_sheet = pd.pivot_table(self.list_spill, index=["День", "Канал"], values="Прокачка, т",
                                             aggfunc="sum").round(1)
        self.list_current_plan = self.list_current_plan.join(self._service_sheet, on=["День", "Канал"])
    def anomaly_search(self):

        logging.info('START search for gas stations with anomalies. The sheet is used "list_spill" ')
        #отбор АЗС с аномалиями
        self._service_sheet = pd.pivot_table(self.list_spill,
                                             index=["День"],
                                             columns=["АЗС"],
                                             values="Прокачка, т",
                                             aggfunc="sum",
                                             fill_value=0).round(1)
        date_count = self._date_report
        average_sheet = pd.DataFrame()

        for count in range(4):
            date_count -= datetime.timedelta(days=7)
            average_sheet = average_sheet.append(self._service_sheet.loc['{}'.format(date_count)], ignore_index=True)

        average_sheet = average_sheet.mean()
        pumping_date_report = self._service_sheet.loc['{}'.format(self._date_report)]
        average_sheet = pd.concat([average_sheet, pumping_date_report], axis=1)
        average_sheet.columns = ['Средняя', 'Прокачка']
        average_sheet['Дельта'] = average_sheet['Прокачка'] - average_sheet['Средняя']
        average_sheet['Дельта%'] = (average_sheet['Прокачка'] - average_sheet['Средняя']) / average_sheet['Прокачка']
        attention_sheet = average_sheet[((average_sheet['Дельта%'] >= 0.3) & (average_sheet['Дельта'] >= 1.5)) |
                                        ((average_sheet['Дельта%'] < -0.3) & (average_sheet['Дельта'] < -0.3)) |
                                        (average_sheet['Прокачка'] == 0)]
        attention_sheet = attention_sheet.drop_duplicates().sort_values('Дельта')
        logging.info('END search for gas stations with anomalies. The sheet is used "list_spill" ')

        logging.info('START receiving deltas through sales channels. The sheet is used "list_spill" ')
        #получение дельт по каналам продаж
        self._service_sheet = pd.pivot_table(self.list_spill,
                                             index=["День"],
                                             columns=["АЗС",  "Канал 2ФЛ, ЮЛ v2"],
                                             values="Прокачка, т",
                                             aggfunc="sum",
                                             fill_value=0).round(1)

        date_count = self._date_report
        channel_average = pd.DataFrame()
        for count in range(4):
            date_count -= datetime.timedelta(days=7)
            channel_average = channel_average.append( self._service_sheet.loc['{}'.format(date_count)], ignore_index=True)
        channel_average = channel_average.mean()
        pumping_date_report = self._service_sheet.loc['{}'.format(self._date_report)]
        channel_average = pd.concat([channel_average, pumping_date_report], axis=1)
        channel_average.columns = ['mean_channel', 'pumping_channel']
        channel_average['delta_channel'] = channel_average['pumping_channel'] - channel_average['mean_channel']
        channel_average = channel_average.drop(columns=['mean_channel', 'pumping_channel'])
        channel_average = channel_average.unstack()
        channel_average.columns = ["Прочее", "ФЛ ГПН", "ФЛ ОПТИ", "ЮЛ ГПН", "ЮЛ Партнера"]
        logging.info('END receiving deltas through sales channels. The sheet is used "list_spill" ')

        logging.info('START search for an oil product that is not being sold. The sheet is used "list_spill" ')
        #ищем какой вид НП не реализуется
        self._service_sheet = pd.pivot_table(self.list_spill[(self.list_spill['День'] >
                                                             pd.to_datetime(self._date_report - datetime.timedelta(days=4))) &
                                             (self.list_spill['День'] <= pd.to_datetime(self._date_report))],
                                             index=["АЗС", "Нефтепродукт"],
                                             columns="День",
                                             values="Прокачка, т",
                                             aggfunc="sum").reset_index().fillna(value=0)
        self._service_sheet = self._service_sheet.query("Нефтепродукт != ['AdBlue', 'АИ-100', 'АИ-98']")
        self._service_sheet.columns = ['АЗС', 'Отсутсвующий НП', '1', '2', '3', 'Прокачка']
        self._service_sheet = self._service_sheet[self._service_sheet['Прокачка'] == 0]
        self._service_sheet = self._service_sheet.drop(columns=['1', '2', '3', 'Прокачка'])
        self._service_sheet = self._service_sheet.drop_duplicates(subset='АЗС', keep=False)
        self._service_sheet = self._service_sheet.set_index('АЗС')
        attention_sheet = pd.concat([attention_sheet, channel_average], axis=1, join="inner")
        attention_sheet = pd.concat([attention_sheet, self._service_sheet], axis=1)
        logging.info('END search for an oil product that is not being sold. The sheet is used "list_spill" ')

        # ищем простои
        azsnumber_previous = None
        interval_alias_previous = None
        nothing_to_do = {'АЗС': [], 'Начало': [], 'Конец': [], 'Длительность': []}
        self.report = self.list_sales_for_10_minutes
        self.report.columns = self.report.iloc[0]
        self.report = self.report.drop(index=0)
        self.report['Интервал'] = pd.to_datetime(self.report['Интервал'])
        self._service_sheet = pd.pivot_table(self.report[self.report['Интервал'].dt.day == self._date_report.day],
                                             index=['АЗС', 'Интервал'],
                                             values='Пролив, т', aggfunc='sum').reset_index().round(2)
        for index, row in self._service_sheet.iterrows():
            if row['АЗС'] == azsnumber_previous:
                if (datetime.time(hour=6, minute=40) < row['Интервал'].time() < datetime.time(hour=20, minute=30)) and \
                        (datetime.time(hour=6, minute=40) < interval_alias_previous.time() < datetime.time(hour=20,
                                                                                                           minute=30)):
                    delta_time = row['Интервал'] - interval_alias_previous
                    if delta_time > datetime.timedelta(minutes=120):
                        nothing_to_do['АЗС'].append(row['АЗС'])
                        nothing_to_do['Начало'].append(interval_alias_previous)
                        nothing_to_do['Конец'].append(row['Интервал'])
                        nothing_to_do['Длительность'].append(delta_time)
            azsnumber_previous = row['АЗС']
            interval_alias_previous = row['Интервал']

        nothing_to_do = pd.DataFrame(data=nothing_to_do)
        nothing_to_do = nothing_to_do.drop_duplicates(subset='АЗС')
        nothing_to_do = nothing_to_do.astype({'АЗС': int})
        nothing_to_do = nothing_to_do.set_index('АЗС')
        attention_sheet = pd.concat([attention_sheet, nothing_to_do], axis=1)

        #ищем изменение цены на НП
        self._service_sheet = pd.pivot_table(self.list_price[self.list_price['Дата'] >=
                                                             pd.to_datetime(self._date_report - datetime.timedelta(days=1))],
                                             index=['АЗС', 'Вид НП'],
                                             columns='Дата',
                                             values='Цена',
                                             aggfunc='sum')
        self._service_sheet = self._service_sheet[f'{self._date_report}'] -\
                              self._service_sheet[f'{self._date_report - datetime.timedelta(days=1)}']
        self._service_sheet = self._service_sheet.loc[lambda x: x != 0]
        self._service_sheet = self._service_sheet.unstack()

        attention_sheet = pd.concat([attention_sheet, self._service_sheet], axis=1).round(2)

        super().save_report(report=attention_sheet, name_sheet='Attention', startrow=1, mode='w')

    def build_and_save(self):

        print('Записываю данные на лист "ОПТИ"')

        # Создание таблицы "Реализация предыдущего месяца"
        if self._date_report.month == 1:
            self.report = pd.pivot_table(
                self.list_current_plan[self.list_current_plan['День'].dt.month == 12],
                index="Канал",
                values=["ТП, тн", "БП, тн", "Прокачка, т"],
                aggfunc="sum",
                margins=True,
                margins_name="Итого").round()
            self.report = self.report[["ТП, тн", "БП, тн", "Прокачка, т"]]
        else:
            self.report = pd.pivot_table(
                self.list_current_plan[self.list_current_plan['День'].dt.month == self._date_report.month - 1],
                index="Канал",
                values=["ТП, тн", "БП, тн", "Прокачка, т"],
                aggfunc="sum",
                margins=True,
                margins_name="Итого").round()
            self.report = self.report[["ТП, тн", "БП, тн", "Прокачка, т"]]
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=3)

        # Создание таблицы "Выполнение предыдущего месяца"
        self.report['ТП,%'] = ((self.report["Прокачка, т"] / self.report["ТП, тн"]) * 100).round()
        self.report['БП,%'] = ((self.report["Прокачка, т"] / self.report["БП, тн"]) * 100).round()
        self.report = self.report[["Прокачка, т", "ТП,%", "БП,%"]]
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=3, startcol=5, overlay="overlay",
                            print_info=0)

        # Создание таблицы "Распределение топливной корзины за предыдущий месяц"
        if self._date_report.month == 1:
            self.report = pd.pivot_table(self.list_spill[self.list_spill['День'].dt.month == 12],
                                         index="Нефтепродукт",
                                         values="Прокачка, т",
                                         aggfunc="sum",
                                         margins=True,
                                         margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
        else:
            self.report = pd.pivot_table(
                self.list_spill[self.list_spill['День'].dt.month == self._date_report.month - 1],
                index="Нефтепродукт",
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=3, startcol=10, overlay="overlay",
                            print_info=0)

        # Создание таблицы "Реализация  предыдущий год, тн"
        self.report = pd.pivot_table(self.list_last_year,
                                     index="Канал",
                                     values="Прокачка, т",
                                     aggfunc="sum",
                                     margins=True,
                                     margins_name="Итого").round()
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=17, overlay="overlay", print_info=0)

        # Создание таблицы "Распределение топливной корзины за предыдущий год"
        self.report = pd.pivot_table(self.list_last_year,
                                     index="Нефтепродукт",
                                     values="Прокачка, т",
                                     aggfunc="sum",
                                     margins=True,
                                     margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=17, startcol=3, overlay="overlay",
                            print_info=0)

        # Создание таблицы "Распределение топливной корзины по сопоставимой сети"
        if self._date_report.day <= 28:
            report_current_year = pd.pivot_table(self.list_spill[(self.list_spill['День'].dt.day <= self._date_report.day) &
                                                                 ((self.list_spill['День'].dt.month == self._date_report.month))],
                                                 index=["АЗС", "Нефтепродукт"],
                                                 values="Прокачка, т",
                                                 aggfunc="sum",
                                                 margins=True,
                                                 margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
            self.report = pd.pivot_table(self.list_last_year[self.list_last_year['Дата'].dt.day <= self._date_report.day],
                                         index=["АЗС", "Нефтепродукт"],
                                         values="Прокачка, т",
                                         aggfunc="sum",
                                         margins=True,
                                         margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)

        else:
            report_current_year = pd.pivot_table(self.list_spill[(self.list_spill['День'].dt.month == self._date_report.month)],
                                                 index=["АЗС", "Нефтепродукт"],
                                                 values="Прокачка, т",
                                                 aggfunc="sum",
                                                 margins=True,
                                                 margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
            self.report = pd.pivot_table(self.list_last_year[self.list_last_year['Дата'].dt.month == self._date_report.month],
                                         index=["АЗС", "Нефтепродукт"],
                                         values="Прокачка, т",
                                         aggfunc="sum",
                                         margins=True,
                                         margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
        self.report = pd.concat([self.report, report_current_year], axis=1, join='inner').reset_index()
        self.report.columns = ["АЗС", "Нефтепродукт", f"{self._date_report.year-1}",  f"{self._date_report.year}"]
        self.report = self.report.groupby('Нефтепродукт', dropna=True).sum()

        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=17, startcol=9, overlay="overlay", print_info=0)

        # Создание таблицы "Реализация на последний день работы"
        self.report = pd.pivot_table(self.list_current_plan[self.list_current_plan['День'] == str(self._date_report)],
                                     index="Канал",
                                     values=["ТП, тн", "БП, тн", "Прокачка, т"],
                                     aggfunc="sum",
                                     margins=True,
                                     margins_name="Итого").round()
        self.report['ТП,%'] = ((self.report["Прокачка, т"] / self.report["ТП, тн"]) * 100).round()
        self.report['БП,%'] = ((self.report["Прокачка, т"] / self.report["БП, тн"]) * 100).round()
        self.report = self.report[["Прокачка, т", "ТП, тн", "ТП,%", "БП, тн", "БП,%"]]
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=30, overlay="overlay", print_info=0)

        # Создание таблицы "Факт+ТП+БП текущий месяц"
        self.report = pd.pivot_table(
            self.list_current_plan[self.list_current_plan['День'].dt.month == self._date_report.month],
            index="Канал",
            values=["ТП, тн", "БП, тн", "Прокачка, т"],
            aggfunc="sum",
            dropna=False,
            margins=True,
            margins_name="Итого").round()
        self.report['ТП,%'] = ((self.report["Прокачка, т"] / self.report["ТП, тн"]) * 100).round()
        self.report['БП,%'] = ((self.report["Прокачка, т"] / self.report["БП, тн"]) * 100).round()
        self.report = self.report[["Прокачка, т", "ТП, тн", "ТП,%", "БП, тн", "БП,%"]]
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=30, startcol=7, overlay="overlay",
                            print_info=0)

        # Создание таблицы "Прогноз на конец месяца (линейно)"
        self.report = self.report[["Прокачка, т", "ТП, тн", "БП, тн"]]
        self.report['Прогноз, т'] = (
                (self.report['Прокачка, т'] / self._date_report.day) * self._date_report_last_day_month).round(1)
        self.report['ТП,%'] = ((self.report["Прогноз, т"] / self.report["ТП, тн"]) * 100).round()
        self.report['БП,%'] = ((self.report["Прогноз, т"] / self.report["БП, тн"]) * 100).round()
        self.report = self.report[["Прогноз, т", "ТП, тн", "ТП,%", "БП, тн", "БП,%"]]
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=30, startcol=14, overlay="overlay",
                            print_info=0)

        # Создание таблицы "Кол-во качающих, АЗС"
        self.pumping_gas_stations = self.list_spill[self.list_spill['День'].dt.month == self._date_report.month][
            'АЗС'].nunique()
        self._service_sheet = self.list_number_of_azs
        self._service_sheet = self.list_number_of_azs.set_index('Дата')

        if self._service_sheet.loc['{}-{}-01'.format(self._date_report.year, self._date_report.month)][
            'Кол-во качающих, АЗС'] != True:
            self._service_sheet.loc['{}-{}-01'.format(self._date_report.year, self._date_report.month)] = \
                [self._date_report.year, self._date_report.month, self.pumping_gas_stations]
        elif self._service_sheet.loc['{}-{}-01'.format(self._date_report.year, self._date_report.month)][
            'Кол-во качающих, АЗС'] != \
                self.pumping_gas_stations:
            self._service_sheet.loc['{}-{}-01'.format(self._date_report.year, self._date_report.month)] = \
                [self._date_report.year, self._date_report.month, self.pumping_gas_stations]

        self.report = pd.pivot_table(self._service_sheet,
                                     columns=["Год", "Месяц"],
                                     values="Кол-во качающих, АЗС",
                                     aggfunc="sum")
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=37, overlay="overlay", print_info=0)

        # Создание таблицы "Распределение сегментов продаж на дату"
        self.report = pd.pivot_table(self.list_spill[self.list_spill['День'] == str(self._date_report)],
                                     index="Канал 2ФЛ, ЮЛ v2",
                                     values="Прокачка, т",
                                     aggfunc="sum",
                                     margins=True,
                                     margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=44, overlay="overlay", print_info=0)

        # Создание таблицы "Рааспределение топливной корзины на дату"
        self.report = pd.pivot_table(self.list_spill[self.list_spill['День'] == str(self._date_report)],
                                     index="Нефтепродукт",
                                     values="Прокачка, т",
                                     aggfunc="sum",
                                     margins=True,
                                     margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=44, startcol=9, overlay="overlay",
                            print_info=0)

        # Создание таблицы "Распределение топливной корзины накопительно на месяц отчета
        self.report = pd.pivot_table(self.list_spill[self.list_spill['День'].dt.month == self._date_report.month],
                                     index="Нефтепродукт",
                                     values="Прокачка, т",
                                     aggfunc="sum",
                                     margins=True,
                                     margins_name="Итого").round().sort_values(by="Прокачка, т", ascending=False)
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=44, startcol=18, overlay="overlay",
                            print_info=0)

        # Создание таблицы "ФЛ/ЮЛ по неделям"
        if self._date_report.isocalendar().week <= 5:
            self.report = pd.pivot_table(self.list_spill
                                         [(self.list_spill[
                                               'День'].dt.isocalendar().week >= 47 + self._date_report_week) |
                                          (self.list_spill['День'].dt.isocalendar().week <= 5)],
                                         index="Канал",
                                         columns="НГ",
                                         values="Прокачка, т",
                                         aggfunc="sum").round()

        else:
            self.report = pd.pivot_table(self.list_spill
                                         [self.list_spill['День'].dt.isocalendar().week >= self._date_report_week - 5],
                                         index="Канал",
                                         columns="НГ",
                                         values="Прокачка, т",
                                         aggfunc="sum").round()
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=60, overlay="overlay", print_info=0)

        # Создание таблицы "ФЛ/ЮЛ ГПН/ЮЛ партнера по неделям"
        if self._date_report.isocalendar().week <= 5:
            self.report = pd.pivot_table(self.list_spill
                                         [(self.list_spill[
                                               'День'].dt.isocalendar().week >= 47 + self._date_report_week) |
                                          (self.list_spill['День'].dt.isocalendar().week <= 5)],
                                         index="Канал 3",
                                         columns="НГ",
                                         values="Прокачка, т",
                                         aggfunc="sum").round()
        else:
            self.report = pd.pivot_table(self.list_spill
                                         [self.list_spill['День'].dt.isocalendar().week >= self._date_report_week - 5],
                                         index="Канал 3",
                                         columns="НГ",
                                         values="Прокачка, т",
                                         aggfunc="sum").round()
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=60, startcol=8, overlay="overlay",
                            print_info=0)

        # Создание таблицы "ФЛ ГПН/ФЛ ОПТИ/Прочее  по неделям"
        if self._date_report.isocalendar().week <= 5:
            self.report = pd.pivot_table(
                self.list_spill[((self.list_spill['День'].dt.isocalendar().week >= 47 + self._date_report_week) |
                                 (self.list_spill['День'].dt.isocalendar().week <= 5)) &
                                ((self.list_spill["Канал 2ФЛ, ЮЛ v2"] == "ФЛ ГПН") |
                                 (self.list_spill["Канал 2ФЛ, ЮЛ v2"] == "ФЛ ОПТИ") |
                                 (self.list_spill["Канал 2ФЛ, ЮЛ v2"] == "Прочее"))],
                index="Канал 2ФЛ, ЮЛ v2",
                columns="НГ",
                values="Прокачка, т",
                aggfunc="sum").round()
        else:
            self.report = pd.pivot_table(self.list_spill[(self.list_spill['День'].dt.isocalendar().week >= self._date_report_week - 5) &
                                                         ((self.list_spill["Канал 2ФЛ, ЮЛ v2"] == "ФЛ ГПН") |
                                                          (self.list_spill["Канал 2ФЛ, ЮЛ v2"] == "ФЛ ОПТИ") |
                                                          (self.list_spill["Канал 2ФЛ, ЮЛ v2"] == "Прочее"))],
                                         index="Канал 2ФЛ, ЮЛ v2",
                                         columns="НГ",
                                         values="Прокачка, т",
                                         aggfunc="sum").round()
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=60, startcol=16, overlay="overlay",
                            print_info=0)

        # Создание таблицы "Выполнение ТП, БП на дату, сср" большая внизу
        self.report = pd.pivot_table(
            self.list_current_plan[(self.list_current_plan['День'].dt.month == self._date_report.month) &
                                   (self.list_current_plan['Прокачка, т'] > 0)],
            index="День",
            columns="Канал",
            values=["ТП, тн", "БП, тн", "Прокачка, т"],
            aggfunc="sum",
            margins=True,
            margins_name="Итого").round()
        column_fl_tp_percent = ((self.report["Прокачка, т"]["ФЛ"] / self.report["ТП, тн"]["ФЛ"]) * 100).round()
        column_ul_tp_percent = ((self.report["Прокачка, т"]["ЮЛ"] / self.report["ТП, тн"]["ЮЛ"]) * 100).round()
        column_fl_bp_percent = ((self.report["Прокачка, т"]["ФЛ"] / self.report["БП, тн"]["ФЛ"]) * 100).round()
        column_ul_bp_percent = ((self.report["Прокачка, т"]["ЮЛ"] / self.report["БП, тн"]["ЮЛ"]) * 100).round()
        column_ul_in_total_percent = (
                (self.report["Прокачка, т"]["ЮЛ"] / self.report["Прокачка, т"]["Итого"]) * 100).round()
        column_fl_in_total_percent = (
                (self.report["Прокачка, т"]["ФЛ"] / self.report["Прокачка, т"]["Итого"]) * 100).round()
        column_tp_total_percent = ((self.report["Прокачка, т"]["Итого"] / self.report["ТП, тн"]["Итого"]) * 100).round()
        column_bp_total_percent = ((self.report["Прокачка, т"]["Итого"] / self.report["БП, тн"]["Итого"]) * 100).round()
        column_fact_total = self.report["Прокачка, т"]["Итого"]
        column_fl_fact, column_ul_fact = self.report["Прокачка, т"]["ФЛ"], self.report["Прокачка, т"]["ЮЛ"]
        self.report = pd.concat([column_fl_in_total_percent, column_fl_fact, column_fl_tp_percent, column_fl_bp_percent,
                                 column_ul_in_total_percent, column_ul_fact, column_ul_tp_percent, column_ul_bp_percent,
                                 column_fact_total, column_tp_total_percent, column_bp_total_percent], axis=1)
        self.report.columns = ["Доля продаж ФЛ,%", "Факт ФЛ", "ТП,%", "БП,%",
                               "Доля продаж ЮЛ,%", "Факт ЮЛ", "ТП,%", "БП,%",
                               "Итого факт", "ТП,%", "БП,%"]
        del (column_fl_tp_percent, column_ul_tp_percent, column_fl_bp_percent, column_ul_bp_percent,
             column_ul_in_total_percent, column_fl_in_total_percent, column_tp_total_percent,
             column_bp_total_percent, column_fact_total, column_fl_fact, column_ul_fact)

        self._service_sheet = pd.pivot_table(
            self.list_spill[self.list_spill['День'].dt.month == self._date_report.month],
            index='День',
            values='АЗС',
            aggfunc='nunique')
        self._service_sheet.loc["Итого"] = \
            (self._service_sheet.loc[f"{self._date_report.year}-{self._date_report.month}-01"] + \
             self._service_sheet.loc['{}'.format(self._date_report)]) / 2
        self.report = self.report.join(self._service_sheet, on='День')

        self._service_sheet = pd.pivot_table(
            self.list_receipts[(self.list_receipts['День'].dt.month == self._date_report.month)],
            index='День',
            columns='Канал',
            values=['Чеки, шт', 'Прокачка, л'],
            aggfunc="sum",
            margins=True,
            margins_name='Итого').round()

        column_prol_l_fl = self._service_sheet['Прокачка, л']['Физические лица']
        column_prol_l_ul = self._service_sheet['Прокачка, л']['Юридические лица']
        column_receipts_fl = self._service_sheet['Чеки, шт']['Физические лица']
        column_receipts_ul = self._service_sheet['Чеки, шт']['Юридические лица']
        self._service_sheet = pd.concat([column_prol_l_fl, column_receipts_fl, column_prol_l_ul, column_receipts_ul],
                                        axis=1)

        del (column_prol_l_fl, column_receipts_fl, column_prol_l_ul, column_receipts_ul)

        self._service_sheet.columns = ['Л ФЛ', 'Чеков ФЛ', 'Л ЮЛ', 'Чеков ЮЛ']
        self._service_sheet['Ср. чек ФЛ, л'] = (
                self._service_sheet['Л ФЛ'] / self._service_sheet['Чеков ФЛ']).round(1)
        self._service_sheet['Ср. чек ЮЛ, л'] = (
                self._service_sheet['Л ЮЛ'] / self._service_sheet['Чеков ЮЛ']).round(1)
        self.report = self.report.join(self._service_sheet, on='День')
        self.report['ССР ФЛ'] = (self.report['Факт ФЛ'] / self.report['АЗС']).round(1)
        self.report['CCР ЮЛ'] = (self.report['Факт ЮЛ'] / self.report['АЗС']).round(1)
        self.report['CCР общ'] = (self.report['Итого факт'] / self.report['АЗС']).round(1)
        self.report = self.report.drop(columns=['Л ФЛ', 'Л ЮЛ'])
        # self.report = self.report[['Доля продаж ФЛ,%', 'Факт ФЛ ТП,%', 'БП,%', 'Доля продаж ЮЛ,%',
        #                           'Факт ЮЛ ТП,%', 'БП,%', 'Итого факт', 'ТП,%', 'БП,%', 'АЗС',
        #                           'ССР ФЛ', 'CCР ЮЛ', 'CCР общ', 'Кол-во чеков ФЛ',
        #                           'Кол-во чеков ЮЛ', 'Сред чек ФЛ, л', 'Сред чек ЮЛ, л']]
        super().save_report(report=self.report, name_sheet='ОПТИ', startrow=84, overlay="overlay", print_info=0)

        print('Записал все данные на лист "ОПТИ"')
        print('______________________________________________________________')

        self.report = self.list_spill.groupby(['День', 'Партнер', 'АЗС', 'НГ', 'Год ввода'], dropna=True).sum()
        self.report = self.report.reset_index()
        if self._date_report.isocalendar().week <= 5:
            self.report = pd.pivot_table(
                self.report[(self.report['День'].dt.isocalendar().week >= 48 + self._date_report_week) |
                            (self.report['День'].dt.isocalendar().week <= 5)],
                index=["Партнер", "Год ввода", "АЗС"],
                columns="НГ",
                values="Прокачка, т",
                aggfunc="mean",
                margins=True,
                margins_name="Итого").reset_index().round(2)
        else:
            self.report = pd.pivot_table(
                self.report[self.report['День'].dt.isocalendar().week > self._date_report_week - 5],
                index=["Партнер", "Год ввода", "АЗС"],
                columns="НГ",
                values="Прокачка, т",
                aggfunc="mean",
                margins=True,
                margins_name="Итого").reset_index().round(2)

        super().save_report(report=self.report, startrow=9, name_sheet='Рейтинг')

        if self._date_report.isocalendar().week <= 3:
            self.report = pd.pivot_table(
                self.list_spill[(self.list_spill['День'].dt.isocalendar().week >= 50 + self._date_report_week) |
                                (self.list_spill['День'].dt.isocalendar().week <= 4)],
                index=["День"],
                columns=["Партнер", "АЗС"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        else:
            self.report = pd.pivot_table(
                self.list_spill[(self.list_spill['День'].dt.isocalendar().week > self._date_report_week - 3) &
                                (self.list_spill['День'].dt.isocalendar().year == self._date_report.year)],
                index=["День"],
                columns=["Партнер", "АЗС"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        super().save_report(report=self.report, name_sheet='Сводная по АЗС')

        if self._date_report.isocalendar().week <= 3:
            self.report = pd.pivot_table(
                self.list_spill[(self.list_spill['День'].dt.isocalendar().week >= 50 + self._date_report_week) |
                                (self.list_spill['День'].dt.isocalendar().week <= 4)],
                index=["День"],
                columns=["Партнер", "АЗС", "Канал"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        else:
            self.report = pd.pivot_table(
                self.list_spill[(self.list_spill['День'].dt.isocalendar().week > self._date_report_week - 3) &
                                (self.list_spill['День'].dt.isocalendar().year == self._date_report.year)],
                index=["День"],
                columns=["Партнер", "АЗС", "Канал"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        super().save_report(report=self.report, name_sheet='Сводная по АЗС(ФЛ, ЮЛ)')

        if self._date_report.isocalendar().week <= 3:
            self.report = pd.pivot_table(
                self.list_spill[((self.list_spill['День'].dt.isocalendar().week >= 50 + self._date_report_week) |
                                 (self.list_spill['День'].dt.isocalendar().week <= 4)) &
                                (self.list_spill['Канал'] == 'ЮЛ')],
                index=["День"],
                columns=["Партнер", "АЗС", "Канал 2ФЛ, ЮЛ v2"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        else:
            self.report = pd.pivot_table(
                self.list_spill[(self.list_spill['День'].dt.isocalendar().week > self._date_report_week - 3) &
                                (self.list_spill['День'].dt.isocalendar().year == self._date_report.year) &
                                (self.list_spill['Канал'] == 'ЮЛ')],
                index=["День"],
                columns=["Партнер", "АЗС", "Канал 2ФЛ, ЮЛ v2"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        super().save_report(report=self.report, name_sheet='ЮЛ(ГПН, Партнер)')

        if self._date_report.isocalendar().week <= 3:
            self.report = pd.pivot_table(
                self.list_spill[((self.list_spill['День'].dt.isocalendar().week >= 50 + self._date_report_week) |
                                 (self.list_spill['День'].dt.isocalendar().week <= 4)) &
                                (self.list_spill['Канал'] == 'ФЛ')],
                index=["День"],
                columns=["Партнер", "АЗС", "Канал 2ФЛ, ЮЛ v2"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        else:
            self.report = pd.pivot_table(
                self.list_spill[(self.list_spill['День'].dt.isocalendar().week > self._date_report_week - 3) &
                                (self.list_spill['День'].dt.isocalendar().year == self._date_report.year) &
                                (self.list_spill['Канал'] == 'ФЛ')],
                index=["День"],
                columns=["Партнер", "АЗС", "Канал 2ФЛ, ЮЛ v2"],
                values="Прокачка, т",
                aggfunc="sum",
                margins=True,
                margins_name="Итого").reset_index().round(1)
        super().save_report(report=self.report, name_sheet='ФЛ(ГПН,ОПТИ,пр.)')

        self.report = pd.pivot_table(
            self.list_spill[self.list_spill['День'] >= pd.to_datetime(self._date_report - datetime.timedelta(days=14))],
            index=["Партнер", "АЗС", "Нефтепродукт"],
            columns="День",
            values="Прокачка, т",
            aggfunc="sum",
            margins=True,
            margins_name="Итого").reset_index().round(1)
        super().save_report(report=self.report, name_sheet='Вид НП по АЗС')

        # Сюда нужно добавить создание листов светофор

        self.report = pd.pivot_table(
            self.list_price[self.list_price['Дата'] >= pd.to_datetime(self._date_report - datetime.timedelta(days=14))],
            index=["Партнер", "АЗС", "Вид НП"],
            columns=["Дата"],
            values="Цена",
            aggfunc="sum").reset_index().round(2)
        super().save_report(report=self.report, name_sheet='Цена стелы')

        print("Отчет собран и сохранен")

    def decoration(self, color='DDEBF7'):

        def formatting_pivot_tables(idx1, idx2, worksheet, freeze_panes):
            max_column_letter = worksheet.cell(row=idx1-7, column=worksheet.max_column).column_letter
            worksheet.freeze_panes = freeze_panes

            worksheet.insert_rows(idx=idx1)
            worksheet.insert_rows(idx=idx2)
            worksheet.insert_rows(idx=worksheet.max_row)
            worksheet[f'B{idx1}'] = 'Итого'
            worksheet[f'B{idx2}'] = 'Итого'
            worksheet[f'B{worksheet.max_row - 1}'] = 'Итого'
            worksheet.column_dimensions['B'].width = 10

            # вставляем формулы
            for row in worksheet[f'C{idx1}:{max_column_letter}{idx1}']:
                for cell in row:
                    cell.value = f'=SUM({cell.column_letter}{idx1 - 7}:{cell.column_letter}{idx1 - 1})'
            for row in worksheet[f'C{idx2}:{max_column_letter}{idx2}']:
                for cell in row:
                    cell.value = f'=SUM({cell.column_letter}{idx2 - 7}:{cell.column_letter}{idx2 - 1})'
            for row in worksheet[f'C{worksheet.max_row - 1}:{max_column_letter}{worksheet.max_row - 1}']:
                for cell in row:
                    cell.value = f'=SUM({cell.column_letter}{idx2 + 1}:{cell.column_letter}{worksheet.max_row - 2})'

            for row in worksheet[f'B1:{max_column_letter}{idx1-8}']:
                for cell in row:
                    cell.fill = PatternFill('solid', start_color=color)

            for row in worksheet[f'B{worksheet.max_row}:{max_column_letter}{ worksheet.max_row}']:
                for cell in row:
                    cell.fill = PatternFill('solid', start_color=color)

            for row in worksheet[f'B{idx1-7}:B{worksheet.max_row}']:
                for cell in row:
                    cell.number_format = 'DD.MMM DDD'

            # форматируем для каждого промежутка отдельно, чтобы не зацепить итоги
            for rows in worksheet.iter_rows(min_col=3, max_col=worksheet.max_column,
                                            max_row=4, min_row=4):
                for cell in rows:
                    formating_string = f'{cell.column_letter}{idx1 - 7}:{cell.column_letter}{idx1 - 1} ' \
                                       f'{cell.column_letter}{idx2 - 7}:{cell.column_letter}{idx2 - 1} ' \
                                       f'{cell.column_letter}{idx2 + 1}:{cell.column_letter}{worksheet.max_row - 2}'

                    worksheet.conditional_formatting.add(formating_string, rule1)

            for row in worksheet['{}1:{}{}'.format(max_column_letter, max_column_letter, worksheet.max_row)]:
                for cell in row:
                    cell.border = Border(left=Side(style='thin'))

            # worksheet.unmerge_cells(start_row=idx1-7, start_column=1, end_row=worksheet.max_row-2, end_column=1)
            # worksheet.merge_cells(f'A5:A{idx1}')
            # worksheet.merge_cells(f'A{idx2-7}:A{idx2}')
            # worksheet.merge_cells(f'A{idx2+1}:A{worksheet.max_row-1}')

        print('______________________________________________________________')
        print('Начинаю оформление отчета')
        self.report = xl.load_workbook(
            '{}/{}_{}.xlsx'.format(self._report_book_directory, self.report_name, self._date_report))

        # создали условное форматирование
        rule = IconSetRule('3TrafficLights1', type='num', values=[0, 90, 100])
        rule1 = ColorScaleRule(start_type='percentile', start_value=0, start_color='F8696B',
                               mid_type='percentile', mid_value=50, mid_color='FFEB84',
                               end_type='percentile', end_value=100, end_color='63BE7B')
        rule2 = IconSetRule('3TrafficLights1', type='percent', values=[0, 33, 67])

        worksheet = self.report['Attention']

        worksheet.column_dimensions['M'].width = 16
        worksheet.column_dimensions['L'].width = 13
        worksheet.column_dimensions['K'].width = 13
        worksheet.column_dimensions['P'].width = 13

        worksheet.insert_cols(idx=11, amount=2)
        worksheet['K2'] = 'Канал'
        worksheet['L2'] = 'Дельта по каналу'
        worksheet.merge_cells('B1:E1')
        worksheet.merge_cells('F1:J1')
        worksheet.merge_cells('K1:L1')
        worksheet.merge_cells('N1:P1')
        worksheet.merge_cells(f'Q1:{get_column_letter(worksheet.max_column)}1')
        worksheet['B1'] = 'Прокачка'
        worksheet['F1'] = 'Канал продаж'
        worksheet['K1'] = 'Канал с max дельтой'
        worksheet['N1'] = 'Простой'
        worksheet['Q1'] = 'Дельта цены на НП'
        worksheet.freeze_panes = 'B3'

        for rows in worksheet.iter_rows(min_col=6, max_col=10,
                                        min_row=3, max_row=worksheet.max_row):
            value = 0
            for cell in rows:
                if (str(cell.value) != 'None') and (str(worksheet[f'C{cell.row}'].value) != 'None'):
                    if worksheet[f'D{cell.row}'].value < 0:
                        if cell.value < value:
                            letter_chanel = cell.column_letter
                            value = cell.value
                            number_rows = cell.row
                    else:
                        if cell.value > value:
                            letter_chanel = cell.column_letter
                            value = cell.value
                            number_rows = cell.row
            worksheet['K{}'.format(number_rows)] = worksheet['{}2'.format(letter_chanel)].value
            worksheet['L{}'.format(number_rows)] = value
            worksheet.conditional_formatting.add(f'F{number_rows}:J{number_rows}', rule2)

            for row in worksheet[f'E3:E{worksheet.max_row}']:
                for cell in row:
                    if str(cell.value) != 'None':
                        if str(cell.value) == '-inf':
                            for row in worksheet[f'A{cell.row}:{get_column_letter(worksheet.max_column)}{cell.row}']:
                                for cell in row:
                                    cell.fill = PatternFill('solid', start_color='DA9694')
                        elif cell.value < 0:
                            for row in worksheet[f'A{cell.row}:{get_column_letter(worksheet.max_column)}{cell.row}']:
                                for cell in row:
                                    cell.fill = PatternFill('solid', start_color='F2DCDB')
                        elif cell.value > 0:
                            for row in worksheet[f'A{cell.row}:{get_column_letter(worksheet.max_column)}{cell.row}']:
                                for cell in row:
                                    cell.fill = PatternFill('solid', start_color='EBF1DE')

            for row in worksheet[f'E3:E{worksheet.max_row}']:
                for cell in row:
                    cell.number_format = "0%"

            for row in worksheet[f'A1:{get_column_letter(worksheet.max_column)}2']:
                for cell in row:
                    cell.font = Font(bold=True)

            for row in worksheet[f'B3:D{worksheet.max_row}']:
                for cell in row:
                    cell.number_format = '0.00'

            for row in worksheet[f'F3:L{worksheet.max_row}']:
                for cell in row:
                    cell.number_format = '0.00'

            for row in worksheet[f'O3:{get_column_letter(worksheet.max_column)}{worksheet.max_row}']:
                for cell in row:
                    cell.number_format = '0.00'

            for row in worksheet[f'N3:P{worksheet.max_row}']:
                for cell in row:
                    cell.number_format = '[$-x-systime]ч:мм:сс AM/PM'

            for row in worksheet[f'B1:{get_column_letter(worksheet.max_column)}2']:
                for cell in row:
                    cell.border = Border(top=Side(style='thin'),
                                         left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         bottom=Side(style='thin'))


        worksheet = self.report['ОПТИ']

        worksheet.column_dimensions['A'].width = 14
        worksheet.column_dimensions['D'].width = 14
        worksheet.column_dimensions['G'].width = 10
        worksheet.column_dimensions['J'].width = 13
        worksheet.column_dimensions['K'].width = 14
        worksheet.column_dimensions['P'].width = 11.30
        worksheet.column_dimensions['Q'].width = 11.30
        worksheet.column_dimensions['S'].width = 13
        worksheet.column_dimensions['T'].width = 14

        worksheet['A1'] = 'Реализация предыдущего месяца, тн'
        if self._date_report.month == 1:
            worksheet['A2'] = 'Год - {}'.format(self._date_report.year - 1)
            worksheet['A3'] = 'Месяц - 12'
        else:
            worksheet['A2'] = 'Год - {}'.format(self._date_report.year)
            worksheet['A3'] = 'Месяц - {}'.format(self._date_report.month - 1)
        # Задаем формат типа 1 000
        for row in worksheet['B5:D7']:
            for cell in row:
                cell.number_format = '# ##0'
        # Красим ячейки заголовков
        for row in worksheet['A4:D4']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        # Красим ячейки итогов и делаем их жирными
        for row in worksheet['A7:D7']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        worksheet['F1'] = 'Выполнение предыдущего месяца, %'
        if self._date_report.month == 1:
            worksheet['F2'] = 'Год - {}'.format(self._date_report.year - 1)
            worksheet['F3'] = 'Месяц - 12'
        else:
            worksheet['F2'] = 'Год - {}'.format(self._date_report.year)
            worksheet['F3'] = 'Месяц - {}'.format(self._date_report.month - 1)
        worksheet.conditional_formatting.add("H5:I7", rule)
        for row in worksheet['G5:G7']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['F4:I4']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['F7:I7']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        worksheet['K1'] = 'Распределение топл. корз. за предыдущий месяц'
        if self._date_report.month == 1:
            worksheet['K2'] = 'Год - {}'.format(self._date_report.year - 1)
            worksheet['K3'] = 'Месяц - 12'
        else:
            worksheet['K2'] = 'Год - {}'.format(self._date_report.year)
            worksheet['K3'] = 'Месяц - {}'.format(self._date_report.month - 1)
        for row in worksheet['L5:L15']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['K4:L5']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        # создаем и вставлем график
        labels = []
        data = []
        for row in worksheet['K6:L14']:
            a = 0
            for call in row:
                if str(call.value) != 'None':
                    if a == 0:
                        labels.append(call.value)
                    else:
                        data.append(call.value)
                    a += 1
        plt.pie(data, labels=labels, autopct='%1.1f%%')
        plt.savefig('{}/paint1.png'.format(self.directory_chart), dpi=50)
        img = Image('{}/paint1.png'.format(self.directory_chart))
        worksheet.add_image(img, 'M2')
        plt.clf()
        plt.cla()
        plt.close()

        worksheet['A15'] = 'Реализация предыдущий год'
        worksheet['A16'] = 'Год - {}'.format(self._date_report.year - 1)
        worksheet['A17'] = 'Месяц - {}'.format(self._date_report.month)
        for row in worksheet['B19:B21']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['A18:B18']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['A21:B21']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        worksheet['D15'] = 'Распределение топл. корз. предыдущий год'
        worksheet['D16'] = 'Год - {}'.format(self._date_report.year - 1)
        worksheet['D17'] = 'Месяц - {}'.format(self._date_report.month)
        for row in worksheet['E19:E29']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['D18:E19']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        worksheet['J16'] = 'Распределение топл. корз. на сопоставимой сети'
        worksheet['J17'] = 'Дата - {}-01 -> {}-{} '.format(self._date_report.month,
                                                           self._date_report.month, self._date_report.day)
        worksheet['J19'] = 'Итого'
        worksheet['K19'] = '=SUM(K20:K28)'
        worksheet['L19'] = '=SUM(L20:L28)'

        for row in worksheet['K19:L27']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['J18:L19']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        # labels = [f'{worksheet["K18"].value}', f'{worksheet["L18"].value}']
        # data_last_year = []
        # data_year = []
        # ytics = []
        # count = 0
        # index = arange(5)
        #
        # for row in worksheet['K20:L25']:
        #     count += 1
        #     for cell in row:
        #         if str(cell.value) != 'None':
        #             if count % 2 == 0: data_year.append(cell.value)
        #             else: data_last_year.append(cell.value)
        # for row in worksheet['J20:J25']:
        #     for cell in row:
        #         ytics.append(f'{cell.value}')
        # plt.axis([0, 8, 0, 5])
        # plt.barh(index, data_last_year, 0.5, color='g')
        # plt.barh(index+0.5, data_year, 0.5, color='r')
        # plt.yticks(index+0.6, ytics)
        # plt.savefig('{}/paint200.png'.format(self.directory_chart), dpi=60)
        # img = Image('{}/paint200.png'.format(self.directory_chart))
        # worksheet.add_image(img, 'M17')
        # plt.clf()
        # plt.cla()
        # plt.close()
        # del(data_last_year, data_year, labels, count, ytics, index)


        worksheet['A29'] = 'Реализация на последний день работы'
        worksheet['A30'] = 'Дата - {}'.format(self._date_report)
        worksheet.conditional_formatting.add("D32:D34", rule)
        worksheet.conditional_formatting.add("F32:F34", rule)
        for row in worksheet['B32:F34']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['A31:F31']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['A34:F34']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        worksheet['H29'] = 'С начала месяца (нарастающим итогом)'
        worksheet['H30'] = 'Дата - {}-{}-01 -> {} '.format(self._date_report.year, self._date_report.month,
                                                           self._date_report)
        worksheet.conditional_formatting.add("K32:K34", rule)
        worksheet.conditional_formatting.add("M32:M34", rule)
        for row in worksheet['I32:M34']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['H31:M31']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['H34:M34']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        worksheet['O29'] = 'Прогноз на конец месяца (линейно)'
        worksheet['O30'] = 'Месяц - {}'.format(self._date_report.month)
        worksheet.conditional_formatting.add("R32:R34", rule)
        worksheet.conditional_formatting.add("T32:T34", rule)
        for row in worksheet['P32:T34']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['O31:T31']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['O34:T34']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
                cell.font = Font(bold=True)

        worksheet['A37'] = 'Кол-во качающих АЗС'
        for row in worksheet['A38:A41']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)

        worksheet['A43'] = 'Распределение сегментов продаж'
        worksheet['A44'] = 'Дата - {}'.format(self._date_report)
        for row in worksheet['B46:B54']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['A45:B46']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        labels = []
        data = []
        for row in worksheet['A47:B54']:
            a = 0
            for call in row:
                if str(call.value) != 'None':
                    if a == 0:
                        labels.append(call.value)
                    else:
                        data.append(call.value)
                    a += 1
        plt.pie(data, labels=labels, autopct='%1.1f%%')
        plt.savefig('{}/paint2.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint2.png'.format(self.directory_chart))
        worksheet.add_image(img, 'C44')
        plt.clf()
        plt.cla()
        plt.close()

        worksheet['J43'] = 'Распределение топл. корз.'
        worksheet['J44'] = 'Дата - {}'.format(self._date_report)
        for row in worksheet['K46:K57']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['J45:K46']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        labels = []
        data = []
        for row in worksheet['J47:K57']:
            a = 0
            for call in row:
                if str(call.value) != 'None':
                    if a == 0:
                        labels.append(call.value)
                    else:
                        data.append(call.value)
                    a += 1
        plt.pie(data, labels=labels, autopct='%1.1f%%')
        plt.savefig('{}/paint3.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint3.png'.format(self.directory_chart))
        worksheet.add_image(img, 'L44')
        plt.clf()
        plt.cla()
        plt.close()

        worksheet['S43'] = 'Распределение топл. корз. накопительно'
        worksheet['S44'] = 'Месяц - {}'.format(self._date_report.month)
        for row in worksheet['T46:T57']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['S45:T46']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        labels = []
        data = []
        for row in worksheet['S47:T57']:
            a = 0
            for call in row:
                if str(call.value) != 'None':
                    if a == 0:
                        labels.append(call.value)
                    else:
                        data.append(call.value)
                    a += 1
        plt.pie(data, labels=labels, autopct='%1.1f%%')
        plt.savefig('{}/paint25.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint25.png'.format(self.directory_chart))
        worksheet.add_image(img, 'U44')
        plt.clf()
        plt.cla()
        plt.close()

        worksheet['A60'] = 'Прокачка, т'
        for row in worksheet['B62:G63']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['A61:G61']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        labels = []
        data = []
        a = 0
        for row in worksheet['B61:G62']:
            for call in row:
                if str(call.value) != 'None':
                    if a < 6:
                        labels.append(call.value)
                    else:
                        data.append(call.value)
                    a += 1
        fig, ax = plt.subplots()
        bar_colors = ['xkcd:gold', 'xkcd:orange', 'xkcd:red', 'xkcd:indigo', 'xkcd:blue', 'xkcd:darkgreen']
        ax.set_title('Динамика прокачки НП ФЛ, тн')
        ax.bar(labels, data, color=bar_colors)
        plt.savefig('{}/paint4.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint4.png'.format(self.directory_chart))
        worksheet.add_image(img, 'A65')
        plt.clf()
        plt.cla()
        plt.close()

        labels = []
        data = []
        for row in worksheet['B61:G61']:
            for call in row:
                if str(call.value) != 'None':
                    labels.append(call.value)
        for row in worksheet['B63:G63']:
            for call in row:
                if str(call.value) != 'None':
                    data.append(call.value)
        fig, ax = plt.subplots()
        bar_colors = ['xkcd:gold', 'xkcd:orange', 'xkcd:red', 'xkcd:indigo', 'xkcd:blue', 'xkcd:darkgreen']
        ax.set_title('Динамика прокачки НП ЮЛ, тн')
        ax.bar(labels, data, color=bar_colors)
        plt.savefig('{}/paint5.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint5.png'.format(self.directory_chart))
        worksheet.add_image(img, 'F65')
        plt.clf()
        plt.cla()
        plt.close()

        worksheet['I60'] = 'Прокачка, т'
        for row in worksheet['J62:O64']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['I61:O61']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        labels = []
        data = []
        for row in worksheet['J61:O61']:
            for call in row:
                if str(call.value) != 'None':
                    labels.append(call.value)
        for row in worksheet['J63:O63']:
            for call in row:
                if str(call.value) != 'None':
                    data.append(call.value)
        fig, ax = plt.subplots()
        bar_colors = ['xkcd:gold', 'xkcd:orange', 'xkcd:red', 'xkcd:indigo', 'xkcd:blue', 'xkcd:darkgreen']
        ax.set_title('Динамика прокачки НП ЮЛ ГПН, тн')
        ax.bar(labels, data, color=bar_colors)
        plt.savefig('{}/paint6.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint6.png'.format(self.directory_chart))
        worksheet.add_image(img, 'K65')
        plt.clf()
        plt.cla()
        plt.close()
        data = []
        for row in worksheet['J64:O64']:
            for call in row:
                if str(call.value) != 'None':
                    data.append(call.value)
        fig, ax = plt.subplots()
        bar_colors = ['xkcd:gold', 'xkcd:orange', 'xkcd:red', 'xkcd:indigo', 'xkcd:blue', 'xkcd:darkgreen']
        ax.set_title('Динамика прокачки НП ЮЛ Партнера, тн')
        ax.bar(labels, data, color=bar_colors)
        plt.savefig('{}/paint7.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint7.png'.format(self.directory_chart))
        worksheet.add_image(img, 'P65')
        plt.clf()
        plt.cla()
        plt.close()

        worksheet['Q60'] = 'Прокачка, т'
        for row in worksheet['R62:W64']:
            for cell in row:
                cell.number_format = '# ##0'
        for row in worksheet['Q61:W61']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        labels = []
        data = []
        for row in worksheet['R61:W61']:
            for call in row:
                if str(call.value) != 'None':
                    labels.append(call.value)
        for row in worksheet['R64:W64']:
            for call in row:
                if str(call.value) != 'None':
                    data.append(call.value)
        fig, ax = plt.subplots()
        bar_colors = ['xkcd:gold', 'xkcd:orange', 'xkcd:red', 'xkcd:indigo', 'xkcd:blue', 'xkcd:darkgreen']
        ax.set_title('Динамика прокачки НП ФЛ ОПТИ, тн')
        ax.bar(labels, data, color=bar_colors)
        plt.savefig('{}/paint8.png'.format(self.directory_chart), dpi=60)
        img = Image('{}/paint8.png'.format(self.directory_chart))
        worksheet.add_image(img, 'U65')
        plt.clf()
        plt.cla()
        plt.close()

        worksheet['A84'] = 'Факт, БП, ТП и ССР'
        worksheet.merge_cells('B84:E84')
        worksheet.merge_cells('F84:I84')
        worksheet['B84'] = 'ФЛ'
        worksheet['F84'] = 'ЮЛ'
        worksheet['B84'].fill = PatternFill('solid', start_color='FCD5B4')
        worksheet['F84'].fill = PatternFill('solid', start_color='D8E4BC')
        worksheet['R{}'.format(worksheet.max_row)] = '=AVERAGE(R86:R{})'.format(worksheet.max_row-1)
        worksheet['S{}'.format(worksheet.max_row)] = '=AVERAGE(S86:S{})'.format(worksheet.max_row - 1)
        worksheet['T{}'.format(worksheet.max_row)] = '=AVERAGE(T86:T{})'.format(worksheet.max_row - 1)

        for row in worksheet['R{}:T{}'.format(worksheet.max_row, worksheet.max_row)]:
            for cell in row:
                cell.number_format = '0.0'
        for row in worksheet['A85:T85']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['A{}:T{}'.format(worksheet.max_row, worksheet.max_row)]:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['A85:A{}'.format(worksheet.max_row)]:
            for cell in row:
                cell.number_format = 'DD.MMM DDD'
        for row in worksheet['B86:Q{}'.format(worksheet.max_row)]:
            for cell in row:
                cell.number_format = '# ##0'
        worksheet.conditional_formatting.add("D85:D{}".format(worksheet.max_row), rule)
        worksheet.conditional_formatting.add("E85:E{}".format(worksheet.max_row), rule)
        worksheet.conditional_formatting.add("H85:H{}".format(worksheet.max_row), rule)
        worksheet.conditional_formatting.add("I85:I{}".format(worksheet.max_row), rule)
        worksheet.conditional_formatting.add("K85:K{}".format(worksheet.max_row), rule)
        worksheet.conditional_formatting.add("L85:L{}".format(worksheet.max_row), rule)

        # смена листа
        worksheet = self.report['Рейтинг']

        worksheet.delete_cols(1)
        worksheet.column_dimensions['A'].width = 16.27
        worksheet.column_dimensions['C'].width = 10.27
        worksheet['C1'] = "Год"
        worksheet['C2'] = 'Недели'
        worksheet.merge_cells('D1:H1')
        worksheet['D1'] = self._date_report.year
        worksheet['D1'].alignment = Alignment(horizontal='center')
        worksheet['C3'] = "до 2 тн."
        worksheet['C4'] = 'от 2 до 4 тн.'
        worksheet['C5'] = "от 4 до 6 тн."
        worksheet['C6'] = 'от 6 до 8 тн.'
        worksheet['C7'] = "свыше 8 тн."
        worksheet['C8'] = 'Итого'
        worksheet.auto_filter.ref = 'A10:I10'
        for row1, row2 in zip(worksheet['D2:H2'], worksheet['D10:H10']):
            for cell1, cell2 in zip(row1, row2):
                cell1.value = cell2.value
                cell1.alignment = Alignment(horizontal='center')
        spis = ['D', 'E', 'F', 'G', 'H']
        for row in worksheet['D3:H3']:
            for letters, cell in zip(spis, row):
                cell.value = '=COUNTIFS({}11:{}{},"<2")'.format(letters, letters, worksheet.max_row - 1)
        for row in worksheet['D4:H4']:
            for letters, cell in zip(spis, row):
                cell.value = '=COUNTIFS({}11:{}{},">=2",{}11:{}{},"<4")'. \
                    format(letters, letters, worksheet.max_row - 1, letters, letters, worksheet.max_row - 1)
        for row in worksheet['D5:H5']:
            for letters, cell in zip(spis, row):
                cell.value = '=COUNTIFS({}11:{}{},">=4",{}11:{}{},"<6")'. \
                    format(letters, letters, worksheet.max_row - 1, letters, letters, worksheet.max_row - 1)
        for row in worksheet['D6:H6']:
            for letters, cell in zip(spis, row):
                cell.value = '=COUNTIFS({}11:{}{},">=6",{}11:{}{},"<8")'. \
                    format(letters, letters, worksheet.max_row - 1, letters, letters, worksheet.max_row - 1)
        for row in worksheet['D7:H7']:
            for letters, cell in zip(spis, row):
                cell.value = '=COUNTIFS({}11:{}{},">=8")'.format(letters, letters, worksheet.max_row - 1)
        for row in worksheet['D8:H8']:
            for letters, cell in zip(spis, row):
                cell.value = '=SUM({}3:{}7)'.format(letters, letters, worksheet.max_row - 1)

        # разукрашиваем верхнюю таблицу
        for row in worksheet['C1:H2']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['C3:H3']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color='FF0000')
        for row in worksheet['C4:H4']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color='FFC000')
        for row in worksheet['C5:H5']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color='FFFF00')
        for row in worksheet['C6:H6']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color='92D050')
        for row in worksheet['C7:H7']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color='00B050')
        for row in worksheet['C8:H8']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)
        for row in worksheet['A10:I10']:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)

        # разукрашиваем ячейки внизу
        for row in worksheet['D11:I{}'.format(worksheet.max_row - 1)]:
            for cell in row:
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center')
                if cell.value != None:
                    if cell.value < 2:
                        cell.fill = PatternFill('solid', start_color='FF0000')
                    elif 2 <= cell.value < 4:
                        cell.fill = PatternFill('solid', start_color='FFC000')
                    elif 4 <= cell.value < 6:
                        cell.fill = PatternFill('solid', start_color='FFFF00')
                    elif 6 <= cell.value < 8:
                        cell.fill = PatternFill('solid', start_color='92D050')
                    elif cell.value >= 8:
                        cell.fill = PatternFill('solid', start_color='00B050')

        for row in worksheet['I11:I{}'.format(worksheet.max_row)]:
            for cell in row:
                cell.border = Border(left=Side(style='thin'))
        worksheet.freeze_panes = 'D11'

        worksheet = self.report['Сводная по АЗС']
        formatting_pivot_tables(idx1=11, idx2=19, worksheet=worksheet, freeze_panes='C4')

        worksheet = self.report['Сводная по АЗС(ФЛ, ЮЛ)']
        formatting_pivot_tables(idx1=12, idx2=20, worksheet=worksheet, freeze_panes='C5')

        worksheet = self.report['ЮЛ(ГПН, Партнер)']
        formatting_pivot_tables(idx1=12, idx2=20, worksheet=worksheet, freeze_panes='C5')

        worksheet = self.report['ФЛ(ГПН,ОПТИ,пр.)']
        formatting_pivot_tables(idx1=12, idx2=20, worksheet=worksheet, freeze_panes='C5')

        worksheet = self.report['Вид НП по АЗС']
        worksheet.delete_cols(1)
        worksheet.column_dimensions['A'].width = 24
        worksheet.freeze_panes = 'C2'
        max_column_letter = worksheet.cell(row=5, column=worksheet.max_column).column_letter
        worksheet.auto_filter.ref = 'A1:S1'
        for row in worksheet['A1:{}1'.format(max_column_letter)]:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)

        for row in worksheet['D1:{}1'.format(max_column_letter)]:
            for cell in row:
                cell.number_format = 'DD.MMM'

        for row in worksheet['D2:{}{}'.format(max_column_letter, worksheet.max_row - 1)]:
            for cell in row:
                if cell.value == None:
                    cell.fill = PatternFill('solid', start_color='FF0000')

        for row in worksheet['A{}:{}{}'.format(worksheet.max_row,
                                               max_column_letter,
                                               worksheet.max_row)]:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)

        for row in worksheet['{}1:{}{}'.format(max_column_letter, max_column_letter, worksheet.max_row)]:
            for cell in row:
                cell.border = Border(left=Side(style='thin'))

        worksheet = self.report['Цена стелы']
        worksheet.delete_cols(1)
        worksheet.column_dimensions['A'].width = 24
        max_column_letter = worksheet.cell(row=4, column=worksheet.max_column).column_letter
        worksheet.auto_filter.ref = 'A1:R1'
        for row in worksheet['A1:{}1'.format(max_column_letter)]:
            for cell in row:
                cell.fill = PatternFill('solid', start_color=color)

        for row in worksheet['D1:{}1'.format(max_column_letter)]:
            for cell in row:
                cell.number_format = 'DD.MMM'
        worksheet.freeze_panes = 'D2'

        cell_row_previous = worksheet['D2'].row
        cell_value_previous = worksheet['D2'].value
        for rows in worksheet.iter_rows(min_col=4, max_col=worksheet.max_column,
                                        max_row=worksheet.max_row, min_row=2):
            for cell in rows:
                if (cell.row == cell_row_previous):
                    if (cell.value != None) and (cell_value_previous != None):
                        delta_price = cell.value - cell_value_previous
                        if delta_price > 0:
                            cell.fill = PatternFill('solid', start_color='F8696B')
                        elif delta_price < 0:
                            cell.fill = PatternFill('solid', start_color='63BE7B')
                cell_value_previous = cell.value
                cell_row_previous = cell.row

        self.report.save('{}/{}_{}.xlsx'.format(self._report_book_directory, self.report_name, self._date_report))
        print("Отчет оформлен и сохранен")
